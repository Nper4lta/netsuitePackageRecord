import time
import os
import requests
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service

# === CONFIG ===
AFTERSHIP_API_KEY = "asat_3ec8cdf0d5e542da92183e53cb689752" \
""
EXCEL_PATH = "tracking_numbers.xlsx"
OUTPUT_DIR = "screenshots"
CHROMEDRIVER_PATH = "/usr/local/bin/chromedriver"

# === Set up folders ===
os.makedirs(OUTPUT_DIR, exist_ok=True)

# === Set up Selenium ===
service = Service(executable_path=CHROMEDRIVER_PATH)
options = webdriver.ChromeOptions()
options.add_argument("--headless")
options.add_argument("--window-size=1200x800")
driver = webdriver.Chrome(service=service, options=options)

# === Load Excel ===
wb = load_workbook(EXCEL_PATH)
ws = wb.active

# === Process each tracking number ===
# === Process each tracking number ===
for row in ws.iter_rows(min_row=2, max_col=1):
    tracking_cell = row[0]
    tracking_number = str(tracking_cell.value).strip()

    if not tracking_number:
        continue

    print(f"Processing: {tracking_number}")

    # 1. Detect carrier
    detect_url = "https://api.aftership.com/v4/couriers/detect"
    headers = {
        "aftership-api-key": AFTERSHIP_API_KEY,
        "Content-Type": "application/json"
    }
    detect_payload = {"tracking": {"tracking_number": tracking_number}}
    detect_resp = requests.post(detect_url, json=detect_payload, headers=headers)

    try:
        courier = detect_resp.json()["data"]["couriers"][0]["slug"]
    except Exception:
        ws.cell(row=tracking_cell.row, column=2).value = "Carrier not found"
        continue

    print(f"  Carrier: {courier}")

    # 2. Get tracking info
    tracking_url = f"https://api.aftership.com/v4/trackings/{courier}/{tracking_number}"
    track_resp = requests.get(tracking_url, headers=headers)

    if track_resp.status_code == 404:
        # Create tracking if not found
        create_url = "https://api.aftership.com/v4/trackings"
        create_payload = {
            "tracking": {
                "tracking_number": tracking_number,
                "slug": courier
            }
        }
        requests.post(create_url, json=create_payload, headers=headers)
        time.sleep(5)
        track_resp = requests.get(tracking_url, headers=headers)

    try:
        tag = track_resp.json()["data"]["tracking"]["tag"]
        status = track_resp.json()["data"]["tracking"]["subtag_message"]
    except Exception:
        ws.cell(row=tracking_cell.row, column=2).value = "Unable to fetch status"
        continue

    # Record the actual status (e.g. InTransit, Delivered, etc.)
    ws.cell(row=tracking_cell.row, column=2).value = tag
    ws.cell(row=tracking_cell.row, column=3).value = status

    if tag != "Delivered":
        continue

    # 3. Open and screenshot tracking page
    driver.get(f"https://www.aftership.com/track/{courier}/{tracking_number}")
    time.sleep(5)  # Wait for full page to load
    screenshot_path = os.path.join(OUTPUT_DIR, f"{tracking_number}.png")
    driver.save_screenshot(screenshot_path)

    # 4. Save screenshot path
    ws.cell(row=tracking_cell.row, column=4).value = screenshot_path
    print(f"  Screenshot saved: {screenshot_path}")



# Save workbook and quit
wb.save(EXCEL_PATH)
driver.quit()
print("All Done.")

